from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
import os
from typing import List
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==============================
# IMPORTS LOCAIS
# ==============================
import app_parser.pdf_reader as pdf_reader
import app_parser.extrator as extrator
from app_parser.alertas import calcular_alerta

# ==============================
# APP
# ==============================
app = FastAPI(title="Leitor de Pagamentos PDF")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {"request": request}
    )
    
# ==============================
# DIRETÓRIOS
# ==============================
UPLOAD_DIR = "uploads"
DATA_DIR = "data"
EXCEL_PATH = os.path.join(DATA_DIR, "pagamentos.xlsx")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# ==============================
# ROTAS
# ==============================
@app.get("/")
def home():
    return {"status": "API online"}

# ==============================
# UPLOAD E PROCESSAMENTO
# ==============================
@app.post("/upload")
async def upload_pdfs(files: List[UploadFile] = File(...)):
    registros = []

    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            continue

        caminho = os.path.join(UPLOAD_DIR, file.filename)

        with open(caminho, "wb") as f:
            f.write(await file.read())

        texto = pdf_reader.ler_pdf(caminho)
        dados = extrator.extrair_pagamento(texto)

        status, dias = calcular_alerta(dados["vencimento"])

        registros.append({
            "Arquivo": file.filename,
            "Valor": dados["valor"],
            "Vencimento": dados["vencimento"],
            "CNPJ": dados["cnpj"],
            "Status": status,
            "Dias Restantes": dias
        })

    # ==============================
    # GERAR EXCEL
    # ==============================
    df = pd.DataFrame(registros)
    df.to_excel(EXCEL_PATH, index=False)

    # ==============================
    # APLICAR CORES
    # ==============================
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    cores = {
        "OK": "C6EFCE",       # verde
        "ALERTA": "FFEB9C",   # amarelo
        "VENCIDO": "FFC7CE"   # vermelho
    }

    for row in range(2, ws.max_row + 1):
        status = ws[f"E{row}"].value
        cor = cores.get(status)

        if cor:
            fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    wb.save(EXCEL_PATH)

    return registros

# ==============================
# DOWNLOAD DO EXCEL
# ==============================
@app.get("/download-excel")
def download_excel():
    return FileResponse(
        path=EXCEL_PATH,
        filename="pagamentos.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
